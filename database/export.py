"""
Waste Classifier - PowerBI Export Utility
Exports waste classification data into formats optimized for PowerBI:
  - CSV (universal, PowerBI Get Data + Text/CSV)
  - Excel (.xlsx) with multiple sheets (PowerBI Get Data → Excel)
  - JSON (PowerBI Get Data JSON)

The exports are designed with PowerBI in mind:
  - Flat tabular structure (no nesting) for direct import
  - Category probabilities as separate columns for easy measures
  - Pre-computed summary sheet for dashboard KPIS
"""

import csv
import json
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

from database.base import StorageBackend
from database.models import WasteRecord
from utils.logger import get_logger

logger = get_logger(__name__)


class PowerBIExporter:
    """
    Exports waste classification data for PowerBI dashboards.

    Produces analysis-ready files with:
      - Flattened record data (one row per classification)
      - Pre-aggregated summary statistics
      - Category distribution, confidence trends, time series data
    """

    def __init__(self, storage: StorageBackend, output_dir: str = "outputs/exports") -> None:
        """
        Initialize PowerBIExporter.

        Args:
            storage: Connected storage backend.
            output_dir: Directory to write export files.
        """
        self.storage = storage
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)

    #------------------------------------------------------------#
    # CSV Export
    #------------------------------------------------------------#

    def export_csv(
        self,
        filename: str = "waste_classifications.csv",
        records: Optional[List[WasteRecord]] = None,
    )-> str:
        """
        Export records to a flat CSV file ready for PowerBI import.

        Args:
            filename: Output filename.
            records: Specific records to export. Fetches all if None.

        Returns:
            Path to the exported CSV file.
        """
        if records is None:
            records = self.storage.get_all_records()

        if not records:
            logger.warning("No records to export.")
            return ""
        
        filepath = self.output_dir / filename
        flat_records = [r.to_flat_dict() for r in records]

        # Get all column names (union of all records for consistency)
        all_keys = []
        seen = set()
        for rec in flat_records:
            for k in rec.keys():
                if k not in seen:
                    all_keys.append(k)
                    seen.add(k)

        with open(filepath, "w", newline="", encoding="utf-8") as f:
            writer = csv.Dictwriter(f, fieldnames=all_keys) 
            writer.writeheader() 
            writer.writerows(flat_records)

        logger.info("CSV exported: %s (%d records)", filepath, len(records)) 
        return str(filepath)
    
    #------------------------------------------------------------#
    # Excel Export (requires openpyxl)
    #------------------------------------------------------------#

    def export_excel(
        self, 
        filename: str = "waste_classifications.xlsx", 
        records: Optional[List[WasteRecord]] = None, 
    ) -> str:
        """
        Export records to an Excel workbook with multiple sheets: 
          - Sheet 1: 'Classifications'  - All records (flat)
          - Sheet 2: 'Summary'          - Aggregate KPIs
          - Sheet 3: 'Category Dist'    - Category counts & percentages
          - Sheet 4: 'Daily Trend'      - Records per day

        Args:
            filename: Output filename.
            records: Specific records to export. Fetches all if None.

        Returns:
            Path to the exported Excel file.
        """
        try:
            import openpyxl 
            from openpyxl.utils import get_column_letter
        except ImportError:
            logger.warning(
            "openpyxl not installed. Falling back to CSV export. "
            "Install via: pip install openpyxl"
            ) 
            return self.export_csv(filename.replace(".xlsx", ".csv"), records)
        
        if records is None:
            records = self.storage.get_all_records()

        if not records:
            logger.warning("No records to export.") 
            return ""
        
        filepath = self.output_dir / filename 
        wb = openpyxl.Workbook()

        # ============== Sheet 1: Classifications (flat records) ============== 
        ws1 = wb.active
        ws1.title = "Classifications"
        flat_records = [r.to_flat_dict() for r in records]

        all_keys = list(flat_records[0].keys()) if flat_records else []
        ws1.append(all_keys)
        for rec in flat_records:
            ws1.append([rec.get(k, "") for k in all_keys])

        # Auto-fit column widths
        for col_idx, key in enumerate(all_keys, 1):
            col_letter = get_column_letter(col_idx)
            max_len = max(len(str(key)), 12)
            ws1.column_dimensions[col_letter].width = min(max_len + 2, 40)

        # ============== Sheet 2: Summary KPIs ============== 
        ws2 = wb.create_sheet("Summary")
        stats = self.storage.get_summary_stats()

        summary_rows = [
            ["KPI", "Value"],
            ["Total Classifications", stats["total_records"]],
            ["Average Confidence", f"{stats['average_confidence']:.2%}"],
            ["Date Range Start", stats["date_range"].get("earliest", "N/A")], 
            ["Date Range End", stats["date_range"].get("latest", "N/A")],
        ]
        for row in summary_rows:
            ws2.append(row)
        
        ws2.append([])
        ws2.append(["Source", "Count"])
        for src, cnt in stats.get("source_counts", {}).items():
            ws2.append([src, cnt])
        
        ws2.append([])
        ws2.append(["Strategy", "Count"])
        for strat, cnt in stats.get("strategy_counts", ()).items():
            ws2.append([strat, cnt])
        
        # ============== Sheet 3: Category Distribution ============== 
        ws3 = wb.create_sheet("Category Distribution") 
        ws3.append(["category", "Count", "Percentage"])

        total = stats["total_records"] or 1 
        for cat, cnt in stats.get("category_counts", {}).items(): 
            ws3.append([cat, cnt, f"{cnt / total:.1%}"])

        # ============== Sheet 4: Daily Trend ============== 
        ws4 = wb.create_sheet("Daily Trend")
        ws4.append(["Date", "Total", "Avg Confidence"])

        daily = self._compute_daily_trend(records) 
        for date_str, data in sorted(daily.items()): 
            ws4.append([date_str, data["count"], round(data["avg_conf"], 4)])

        wb.save(filepath)
        logger.info("Excel exported: %s (%d records, 4 sheets)", filepath, len(records)) 
        return str(filepath)
    
    #------------------------------------------------------------#
    # JSON Export
    #------------------------------------------------------------#

    def export_json(
        self, 
        filename: str = "waste_classifications.json", 
        records: Optional[List[WasteRecord]] = None,
    ) -> str:
        """
        Export records to a JSON file.

        Args:
            filename: Output filename.
            records: Specific records to export. Fetches all if None.

        Returns:
            Path to the exported JSON file.
        """
        if records is None:
            records = self.storage.get_all_records()
        
        filepath = self.output_dir / filename

        data = {
            "exported_at": datetime.now().isoformat(),
            "total_records": len(records),
            "summary": self.storage.get_summary_stats(),
            "records": [r.to_dict() for r in records],
        }

        with open(filepath, "w", encoding="utf-8") as f: 
            json.dump(data, f, indent=2, ensure_ascii=False, default=str)

        logger.info("JSON exported: %s (%d records)", filepath, len(records))
        return str(filepath)
    
    #------------------------------------------------------------#
    # Export All Formats
    #------------------------------------------------------------#

    def export_all(
        self, records: Optional[List[WasteRecord]] = None
    )-> Dict[str, str]:
        """
        Export in all formats (CSV, Excel, JSON).

        Returns:
            Dict mapping format name to file path.
        """
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

        paths = {
            "csv": self.export_csv(f"waste_data_{timestamp}.csv", records), 
            "excel": self.export_excel(f"waste_data_{timestamp}.xlsx", records), 
            "json": self.export_json(f"waste_data_{timestamp}.json", records),
        }

        logger.info("All exports complete: %s", paths)
        return paths
    
    #------------------------------------------------------------#
    # Helpers
    #------------------------------------------------------------#

    @staticmethod
    def compute_daily_trend(records: List[WasteRecord]) -> Dict[str, Dict]:
        """Aggregate records by date for time-series analysis."""
        daily: Dict[str, Dict] = {}

        for r in records:
            try:
                date_str = r.timestamp[:10] # YYYY-MM-DD
            except (TypeError, IndexError):
                continue

            if date_str not in daily:
                daily[date_str] = {"count": 0, "total_conf": 0.0}

            daily[date_str]["count"] += 1
            daily[date_str]["total_conf"] += r.final_confidence

        # Compute averages
        for date_str in daily:
            cnt = daily[date_str]["count"]
            daily[date_str]["avg_conf"] = daily[date_str]["total_conf"] / cnt if cnt > 0 else 0
            del daily[date_str]["total_conf"]

        return daily